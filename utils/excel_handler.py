"""
utils/excel_handler.py — Excel import/export using openpyxl.

Import
------
* Companies Act sheet: columns Asset Name, Category, Cost, Purchase Date,
  Useful Life, Residual Value %, Method
* Income Tax sheet: columns Block Name, Opening WDV, Additions, Deletions, Rate

Export
------
* Each report type on its own sheet: "Companies Act", "Tax Depreciation", "DTA-DTL"
* Headers are bold, columns are auto-sized, numbers are formatted.
"""

from datetime import date
from typing import List, Optional, Tuple

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _header_style(ws, row: int, col_count: int, fill_colour: str = "2E86AB"):
    """Apply bold font and fill colour to a header row."""
    if not OPENPYXL_AVAILABLE:
        return
    bold_font = Font(bold=True, color="FFFFFF")
    fill = PatternFill(start_color=fill_colour, end_color=fill_colour, fill_type="solid")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = bold_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border


def _auto_width(ws):
    """Adjust column widths based on content."""
    if not OPENPYXL_AVAILABLE:
        return
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _thin_border():
    """Return a thin Border object."""
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


# ---------------------------------------------------------------------------
# IMPORT
# ---------------------------------------------------------------------------

def import_companies_act_data(filepath: str) -> Tuple[List[dict], List[str]]:
    """
    Read the 'Companies Act' sheet from *filepath* and return parsed rows.

    Returns
    -------
    (rows, errors)
        rows  — list of dicts with keys matching AssetInput fields
        errors — list of human-readable error strings
    """
    if not OPENPYXL_AVAILABLE:
        return [], ["openpyxl is not installed. Run: pip install openpyxl"]

    errors: List[str] = []
    rows: List[dict] = []

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as exc:
        return [], [f"Cannot open file: {exc}"]

    sheet_name = None
    for name in wb.sheetnames:
        if "companies" in name.lower() or "act" in name.lower():
            sheet_name = name
            break
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]  # fall back to first sheet

    ws = wb[sheet_name]
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]

    expected = {
        "asset name": "asset_name",
        "category": "category",
        "cost": "cost",
        "purchase date": "purchase_date",
        "useful life": "useful_life",
        "residual value %": "residual_value_pct",
        "method": "method",
    }

    col_map = {}
    for idx, h in enumerate(headers):
        for key, field in expected.items():
            if key in h:
                col_map[field] = idx

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue
        entry = {}
        missing = []
        for field, col_idx in col_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val is None:
                missing.append(field)
            else:
                entry[field] = val
        if missing:
            errors.append(f"Row {row_num}: missing fields {missing}")
            continue
        # Normalise purchase_date
        pd_val = entry.get("purchase_date")
        if isinstance(pd_val, date):
            entry["purchase_date"] = pd_val.strftime("%d/%m/%Y")
        elif pd_val is not None:
            entry["purchase_date"] = str(pd_val)
        rows.append(entry)

    return rows, errors


def import_income_tax_data(filepath: str) -> Tuple[List[dict], List[str]]:
    """
    Read the 'Income Tax' / 'Tax' sheet from *filepath* and return parsed rows.

    Returns
    -------
    (rows, errors)
    """
    if not OPENPYXL_AVAILABLE:
        return [], ["openpyxl is not installed. Run: pip install openpyxl"]

    errors: List[str] = []
    rows: List[dict] = []

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as exc:
        return [], [f"Cannot open file: {exc}"]

    sheet_name = None
    for name in wb.sheetnames:
        if "tax" in name.lower() or "income" in name.lower():
            sheet_name = name
            break
    if sheet_name is None and len(wb.sheetnames) > 1:
        sheet_name = wb.sheetnames[1]
    elif sheet_name is None:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]

    expected = {
        "block name": "block_name",
        "opening wdv": "opening_wdv",
        "additions": "additions",
        "deletions": "deletions",
        "rate": "rate",
    }

    col_map = {}
    for idx, h in enumerate(headers):
        for key, field in expected.items():
            if key in h:
                col_map[field] = idx

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue
        entry = {}
        missing = []
        for field, col_idx in col_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val is None:
                missing.append(field)
            else:
                entry[field] = val
        if missing:
            errors.append(f"Row {row_num}: missing fields {missing}")
            continue
        rows.append(entry)

    return rows, errors


# ---------------------------------------------------------------------------
# EXPORT
# ---------------------------------------------------------------------------

def export_companies_act(ws, schedule_rows: list, asset_name: str):
    """Write Companies Act depreciation schedule to *ws* (a Worksheet)."""
    if not OPENPYXL_AVAILABLE:
        return

    title_row = 1
    ws.cell(row=title_row, column=1, value=f"Asset: {asset_name}")
    ws.cell(row=title_row, column=1).font = Font(bold=True, size=12)

    headers = ["Year", "Opening WDV (₹)", "Depreciation (₹)", "Closing WDV (₹)"]
    header_row = 2
    for col, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=h)
    _header_style(ws, header_row, len(headers))

    border = _thin_border()
    num_fmt = '#,##0.00'
    for r_idx, row in enumerate(schedule_rows, start=header_row + 1):
        ws.cell(row=r_idx, column=1, value=row.year_label).border = border
        for col, val in enumerate([row.opening_wdv, row.depreciation, row.closing_wdv], start=2):
            cell = ws.cell(row=r_idx, column=col, value=val)
            cell.number_format = num_fmt
            cell.border = border

    _auto_width(ws)


def export_income_tax(ws, results: list):
    """Write Income Tax depreciation results to *ws*."""
    if not OPENPYXL_AVAILABLE:
        return

    headers = [
        "Block", "Opening WDV (₹)", "Additions (₹)", "Deletions (₹)",
        "Adjusted WDV (₹)", "Rate (%)", "Depreciation (₹)", "Closing WDV (₹)"
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    _header_style(ws, 1, len(headers))

    border = _thin_border()
    num_fmt = '#,##0.00'
    for r_idx, res in enumerate(results, start=2):
        vals = [
            res.block_name, res.opening_wdv, res.additions, res.deletions,
            res.adjusted_wdv, res.effective_rate, res.depreciation, res.closing_wdv
        ]
        for col, val in enumerate(vals, start=1):
            cell = ws.cell(row=r_idx, column=col, value=val)
            if isinstance(val, float):
                cell.number_format = num_fmt
            cell.border = border

    _auto_width(ws)


def export_dta_dtl(ws, summary):
    """Write DTA/DTL summary to *ws*."""
    if not OPENPYXL_AVAILABLE:
        return

    headers = [
        "Asset", "Book Value (₹)", "Tax Value (₹)",
        "Difference (₹)", "Tax Rate (%)", "DTA (₹)", "DTL (₹)"
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    _header_style(ws, 1, len(headers))

    border = _thin_border()
    num_fmt = '#,##0.00'
    for r_idx, row in enumerate(summary.rows, start=2):
        vals = [
            row.asset_name, row.book_value, row.tax_value,
            row.difference, row.tax_rate, row.dta, row.dtl
        ]
        for col, val in enumerate(vals, start=1):
            cell = ws.cell(row=r_idx, column=col, value=val)
            if isinstance(val, float):
                cell.number_format = num_fmt
            cell.border = border

    # Summary row
    summary_row = len(summary.rows) + 2
    ws.cell(row=summary_row, column=1, value="NET POSITION").font = Font(bold=True)
    ws.cell(row=summary_row, column=6, value=summary.net_dta).number_format = num_fmt
    ws.cell(row=summary_row, column=7, value=summary.net_dtl).number_format = num_fmt
    for col in range(1, 8):
        ws.cell(row=summary_row, column=col).border = border

    _auto_width(ws)


def export_all_to_excel(
    filepath: str,
    ca_schedule_rows: Optional[list] = None,
    ca_asset_name: str = "",
    tax_results: Optional[list] = None,
    dta_summary=None,
) -> Tuple[bool, str]:
    """
    Export all available data to a single Excel workbook with multiple sheets.

    Returns
    -------
    (success, message)
    """
    if not OPENPYXL_AVAILABLE:
        return False, "openpyxl is not installed. Run: pip install openpyxl"

    try:
        wb = openpyxl.Workbook()
        # Remove default sheet
        default_ws = wb.active
        wb.remove(default_ws)

        if ca_schedule_rows:
            ws_ca = wb.create_sheet("Companies Act")
            export_companies_act(ws_ca, ca_schedule_rows, ca_asset_name)

        if tax_results:
            ws_tax = wb.create_sheet("Tax Depreciation")
            export_income_tax(ws_tax, tax_results)

        if dta_summary:
            ws_dta = wb.create_sheet("DTA-DTL")
            export_dta_dtl(ws_dta, dta_summary)

        if not wb.sheetnames:
            return False, "No data to export."

        wb.save(filepath)
        return True, f"Exported successfully to {filepath}"
    except Exception as exc:
        return False, f"Export failed: {exc}"
